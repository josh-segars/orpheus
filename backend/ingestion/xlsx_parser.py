"""LinkedIn Analytics XLSX parser.

Reads the 5-sheet export from linkedin.com/analytics/creator/content/
Sheets: DISCOVERY, ENGAGEMENT, TOP POSTS, FOLLOWERS, DEMOGRAPHICS

Export period is typically "Past 365 days" — the exact range appears
in cell B1 of the DISCOVERY sheet.

Dependencies: openpyxl (in requirements.txt)
"""

import io
import logging
from datetime import datetime

import openpyxl

from backend.ingestion.types import (
    XlsxData,
    DiscoverySummary,
    EngagementRow,
    TopPostItem,
    FollowersData,
    FollowersRow,
    DemographicsData,
)

logger = logging.getLogger(__name__)


def _safe_int(value, default: int = 0) -> int:
    """Convert a cell value to int, handling None and floats."""
    if value is None:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def _safe_float(value, default: float = 0.0) -> float:
    """Convert a cell value to float, handling None."""
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def _safe_str(value, default: str = "") -> str:
    """Convert a cell value to string, handling None."""
    if value is None:
        return default
    return str(value).strip()


def _parse_date_cell(value) -> str:
    """Convert a date cell to a string date.

    LinkedIn uses M/D/YYYY format in the XLSX. openpyxl may return
    a datetime object or a string depending on the cell format.
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    # Try to parse M/D/YYYY string
    s = str(value).strip()
    try:
        dt = datetime.strptime(s, "%m/%d/%Y")
        return dt.strftime("%Y-%m-%d")
    except ValueError:
        return s


def _parse_discovery(ws) -> DiscoverySummary:
    """Parse the DISCOVERY sheet.

    Layout:
        A1: "Overall Performance"  B1: "3/17/2025 - 3/16/2026"
        A2: "Impressions"          B2: 279637
        A3: "Members reached"      B3: 66443
    """
    period = _safe_str(ws.cell(1, 2).value)
    impressions = _safe_int(ws.cell(2, 2).value)
    members_reached = _safe_int(ws.cell(3, 2).value)

    logger.info(
        "DISCOVERY: period=%s, impressions=%d, members_reached=%d",
        period, impressions, members_reached,
    )
    return DiscoverySummary(
        period=period,
        impressions=impressions,
        members_reached=members_reached,
    )


def _parse_engagement(ws) -> list[EngagementRow]:
    """Parse the ENGAGEMENT sheet.

    Layout:
        Row 1: Date | Impressions | Engagements  (header)
        Row 2+: daily data
    """
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rows.append(EngagementRow(
            date=_parse_date_cell(row[0]),
            impressions=_safe_int(row[1]),
            engagements=_safe_int(row[2]) if len(row) > 2 else 0,
        ))

    logger.info("ENGAGEMENT: %d daily rows", len(rows))
    return rows


def _parse_top_posts(ws) -> list[TopPostItem]:
    """Parse the TOP POSTS sheet.

    Layout:
        Row 1: "Maximum of 50 posts available..."  (info text)
        Row 2: blank
        Row 3: Post URL | Post publish date | Engagements | (blank) | Post URL | Post publish date | Impressions | (blank)
        Row 4+: data in two side-by-side rankings

    Left block (cols A-C): top by engagements.
    Right block (cols E-G): top by impressions.
    We merge them into a single list by post URL.
    """
    posts_by_url: dict[str, TopPostItem] = {}

    for row in ws.iter_rows(min_row=4, values_only=True):
        # Left block: top by engagements (cols A, B, C)
        left_url = _safe_str(row[0]) if len(row) > 0 else ""
        if left_url and left_url.startswith("http"):
            if left_url not in posts_by_url:
                posts_by_url[left_url] = TopPostItem(
                    post_url=left_url,
                    published_date=_parse_date_cell(row[1]) if len(row) > 1 else "",
                )
            posts_by_url[left_url].engagements = _safe_int(row[2]) if len(row) > 2 else 0

        # Right block: top by impressions (cols E, F, G)
        right_url = _safe_str(row[4]) if len(row) > 4 else ""
        if right_url and right_url.startswith("http"):
            if right_url not in posts_by_url:
                posts_by_url[right_url] = TopPostItem(
                    post_url=right_url,
                    published_date=_parse_date_cell(row[5]) if len(row) > 5 else "",
                )
            posts_by_url[right_url].impressions = _safe_int(row[6]) if len(row) > 6 else 0

    result = list(posts_by_url.values())
    logger.info("TOP POSTS: %d unique posts merged from two rankings", len(result))
    return result


def _parse_followers(ws) -> FollowersData:
    """Parse the FOLLOWERS sheet.

    Layout:
        Row 1: "Total followers on 3/16/2026:" | 3037
        Row 2: blank
        Row 3: Date | New followers  (header)
        Row 4+: daily data
    """
    total_followers = _safe_int(ws.cell(1, 2).value)

    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if row[0] is None:
            continue
        rows.append(FollowersRow(
            date=_parse_date_cell(row[0]),
            new_followers=_safe_int(row[1]),
        ))

    logger.info(
        "FOLLOWERS: total=%d, %d daily rows",
        total_followers, len(rows),
    )
    return FollowersData(total_followers=total_followers, rows=rows)


def _parse_demographics(ws) -> DemographicsData:
    """Parse the DEMOGRAPHICS sheet.

    Layout:
        Row 1: Top Demographics | Value | Percentage  (header)
        Row 2+: category | value | percentage

    Categories: "Job titles", "Locations", "Industries"
    Percentage is a float (e.g. 0.025 = 2.5%).
    """
    job_titles: dict[str, float] = {}
    locations: dict[str, float] = {}
    industries: dict[str, float] = {}

    category_map = {
        "job titles": job_titles,
        "locations": locations,
        "industries": industries,
    }

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        category = _safe_str(row[0]).lower()
        value = _safe_str(row[1])
        pct = _safe_float(row[2])

        target = category_map.get(category)
        if target is not None and value:
            target[value] = pct

    logger.info(
        "DEMOGRAPHICS: %d job titles, %d locations, %d industries",
        len(job_titles), len(locations), len(industries),
    )
    return DemographicsData(
        job_titles=job_titles,
        locations=locations,
        industries=industries,
    )


def parse_xlsx(xlsx_bytes: bytes) -> XlsxData:
    """Parse a LinkedIn Analytics XLSX export into structured data.

    Args:
        xlsx_bytes: Raw bytes of the XLSX file.

    Returns:
        XlsxData with all 5 sheets parsed.
    """
    wb = openpyxl.load_workbook(
        io.BytesIO(xlsx_bytes),
        data_only=True,
    )

    sheet_names = [s.upper() for s in wb.sheetnames]
    logger.info("XLSX contains %d sheets: %s", len(wb.sheetnames), wb.sheetnames)

    # Map sheet names case-insensitively
    def get_sheet(target: str):
        for name in wb.sheetnames:
            if name.upper() == target.upper():
                return wb[name]
        return None

    discovery = DiscoverySummary()
    engagement = []
    top_posts = []
    followers = FollowersData()
    demographics = DemographicsData()

    ws = get_sheet("DISCOVERY")
    if ws:
        discovery = _parse_discovery(ws)
    else:
        logger.warning("DISCOVERY sheet not found in XLSX")

    ws = get_sheet("ENGAGEMENT")
    if ws:
        engagement = _parse_engagement(ws)
    else:
        logger.warning("ENGAGEMENT sheet not found in XLSX")

    ws = get_sheet("TOP POSTS")
    if ws:
        top_posts = _parse_top_posts(ws)
    else:
        logger.warning("TOP POSTS sheet not found in XLSX")

    ws = get_sheet("FOLLOWERS")
    if ws:
        followers = _parse_followers(ws)
    else:
        logger.warning("FOLLOWERS sheet not found in XLSX")

    ws = get_sheet("DEMOGRAPHICS")
    if ws:
        demographics = _parse_demographics(ws)
    else:
        logger.warning("DEMOGRAPHICS sheet not found in XLSX")

    wb.close()

    result = XlsxData(
        discovery=discovery,
        engagement=engagement,
        top_posts=top_posts,
        followers=followers,
        demographics=demographics,
    )

    logger.info(
        "XLSX parsed — %d engagement days, %d top posts, %d followers total",
        len(engagement), len(top_posts), followers.total_followers,
    )
    return result
